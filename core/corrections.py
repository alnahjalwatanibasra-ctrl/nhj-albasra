# -*- coding: utf-8 -*-
"""تصحيحات المصطلحات، الترقيم التسلسلي، وأدوات الأرقام العربية."""
# openpyxl يُستورد داخل last_book_number فقط — تحميله عند الإقلاع كان يبطئ الفتح 1.4ث

W2A = str.maketrans('0123456789', '٠١٢٣٤٥٦٧٨٩')
A2W = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')


def to_arabic(s):
    return str(s).translate(W2A)


def to_western_digits(s):
    return ''.join(c for c in str(s).translate(A2W) if c.isdigit())


def format_phone(v, arabic=True):
    """يوحّد الهاتف العراقي: ٠٧ + ١١ خانة. يعيد '' إن لم توجد خانات."""
    d = to_western_digits(v)
    if not d:
        return ''
    if len(d) == 10 and d.startswith('7'):
        d = '0' + d
    return to_arabic(d) if arabic else d


def format_ref_date(v, arabic=True):
    """ينسّق تاريخ المرجع. التواريخ المخزّنة datetime أدخلها المستخدم بصيغة يوم/شهر
    لكن Excel فسّرها شهر/يوم (أمريكياً)، فإعادة النص الأصلي = month/day/year."""
    if v is None or str(v).strip() == '':
        return ''
    if hasattr(v, 'year'):
        s = '%d/%d/%d' % (v.month, v.day, v.year)
    else:
        s = str(v).strip()
    return to_arabic(s) if arabic else s


def infer_start_number(nums):
    """يستنتج رقم كتاب البداية من أرقام OCR (السجل متسلسل بطبيعته):
    أغلبية (الرقم - موقع الصف). nums: قائمة سلاسل خانات غربية."""
    from collections import Counter
    votes = Counter()
    for i, d in enumerate(nums):
        if d and d.isdigit() and 2 <= len(d) <= 5:
            votes[int(d) - i] += 1
    if not votes:
        return None
    start, v = votes.most_common(1)[0]
    return start if v >= 2 else None


def apply_replacements(text, replacements):
    if not text:
        return text
    for a, b in replacements.items():
        text = text.replace(a, b)
    return text


def last_book_number(prev_xlsx_path, number_header_hints=('رقم الكتاب', 'رقم')):
    """يقرأ آخر رقم كتاب من ملف سجل سابق (أعلى قيمة رقمية في عمود رقم الكتاب)."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(prev_xlsx_path, data_only=True)
        ws = wb.active
    except Exception:
        return None
    # جد عمود رقم الكتاب
    ncol = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or '')
        if any(hint in h for hint in number_header_hints):
            ncol = c; break
    if not ncol:
        return None
    mx = None
    for r in range(2, ws.max_row + 1):
        d = to_western_digits(ws.cell(r, ncol).value)
        if d.isdigit():
            v = int(d)
            if mx is None or v > mx:
                mx = v
    return mx


def sequential_numbers(count, start):
    """يعيد قائمة أرقام متسلسلة تبدأ من start."""
    return [start + i for i in range(count)]


def sequential_numbers_anchored(ocr_digits, start):
    """ترقيم تسلسلي «مرسّى» بأرقام OCR: إن أسقط القارئ صفاً كاملاً انزاح كل
    ما بعده بالترقيم الأعمى — هنا القفزة تُعتمد إذا أكّدها صفان مقروءان،
    فيبقى كل صف على رقمه الحقيقي وتظهر الفجوة (دليل الصف الساقط) بدل انزياح الكل.
    قراءة شاذة منفردة تُتجاهل. ocr_digits: خانات غربية لكل صف. start: رقم أول صف."""
    n = len(ocr_digits)
    offs = []
    for i, d in enumerate(ocr_digits):
        offs.append(int(d) - i if (d and d.isdigit() and 2 <= len(d) <= 5) else None)
    cur = start
    out = []
    for i in range(n):
        o = offs[i]
        # صف ساقط يزيد الإزاحة بمقدار صغير موجب فقط (+1..+3)؛ الإزاحات الكبيرة
        # أو السالبة أخطاء قراءة منهجية (٤٩٨⟵٤٩١ متتاليةً!) تُتجاهل — درس مُقاس
        if o is not None and 0 < (o - cur) <= 3:
            nxt = next((offs[j] for j in range(i + 1, min(i + 4, n))
                        if offs[j] is not None), None)
            if nxt == o:          # صفّان متتاليان يؤكدان القفزة
                cur = o
        out.append(cur + i)
    return out
