# -*- coding: utf-8 -*-
"""الملف المرجعي: تحميل، كشف الأعمدة تلقائياً، ومطابقة الأسماء لسحب هاتف/معرف/دائرة."""
import unicodedata, difflib
from collections import Counter
import openpyxl

# كلمات مفتاحية لكشف أعمدة المرجع تلقائياً
COL_HINTS = {
    'name':  ['الاسم الرباعي', 'الاسم الثلاثي', 'الاسم', 'اسم صاحب'],
    'phone': ['رقم الهاتف', 'الهاتف', 'الموبايل', 'الجوال'],
    'moar':  ['المعرف', 'المعرّف', 'معرف'],
    'dawira': ['الدائرة', 'دائرة'],
    'subj':  ['الموضوع', 'موضوع'],
    'jiha':  ['الجهة الصادر اليها', 'الجهة المرسل اليها', 'الجهة'],
    'date':  ['تاريخ الطلب', 'تاريخ الكتاب', 'التاريخ'],
    'status': ['الحالة', 'حالة'],
}


def norm(s):
    s = '' if s is None else str(s)
    for a, b in [('أ','ا'),('إ','ا'),('آ','ا'),('ى','ي'),('ة','ه'),('ؤ','و'),('ئ','ي'),('ـ','')]:
        s = s.replace(a, b)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.split()).strip()


def sim(a, b):
    return difflib.SequenceMatcher(None, norm(a), norm(b)).ratio()


def _find_header_row(ws, max_scan=5):
    """يجد صف العناوين (أول صف فيه عدة كلمات مفتاحية معروفة)."""
    best_r, best_score = 1, -1
    for r in range(1, min(max_scan, ws.max_row) + 1):
        vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        score = sum(1 for v in vals for hints in COL_HINTS.values()
                    for h in hints if norm(h) == v)
        if score > best_score:
            best_score, best_r = score, r
    return best_r


class Reference:
    def __init__(self, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        self.header_row = _find_header_row(ws)
        # كشف الأعمدة
        self.colmap = {}
        headers = {c: norm(ws.cell(self.header_row, c).value) for c in range(1, ws.max_column + 1)}
        for role, hints in COL_HINTS.items():
            for c, hv in headers.items():
                if hv and any(norm(h) == hv for h in hints):
                    self.colmap[role] = c; break
        # تحميل الصفوف
        self.rows = []
        if 'name' in self.colmap:
            for r in range(self.header_row + 1, ws.max_row + 1):
                nm = ws.cell(r, self.colmap['name']).value
                if not nm: continue
                self.rows.append({role: ws.cell(r, col).value for role, col in self.colmap.items()})
        self.by_name = [(norm(x.get('name', '')), x) for x in self.rows]
        self.moar_freq = Counter(norm(x.get('moar', '')) for x in self.rows if x.get('moar'))

    def detected_columns(self):
        return {role: 'عمود %d' % col for role, col in self.colmap.items()}

    def vocab(self):
        """قوائم المفردات القياسية (لحقنها في prompt الاستخراج — ترشد قراءة خط اليد)."""
        def distinct(role):
            seen, out = set(), []
            for x in self.rows:
                v = str(x.get(role) or '').strip()
                if v and norm(v) not in seen:
                    seen.add(norm(v)); out.append(v)
            return sorted(out)
        return {
            'أسماء معروفة': distinct('name'),
            'مواضيع معروفة': distinct('subj'),
            'معرفون معروفون': distinct('moar'),
            'جهات معروفة': distinct('jiha'),
        }

    def _match_idx(self, name_candidates, moar='', subj='', th=0.82, ftok_th=0.55, rare_max=4):
        """يطابق الاسم؛ يعيد فهرس صف المرجع المطابق أو None."""
        cands = [c for c in name_candidates if c and len(norm(c)) >= 5]
        if not cands or not self.rows:
            return None
        bidx, bname, brank = None, 0, -1
        for i, x in enumerate(self.rows):
            ns = max(sim(c, x.get('name', '')) for c in cands)
            if ns < 0.66:
                continue
            ms = sim(moar, x.get('moar', '')) if moar else 0
            rare = ms >= 0.8 and self.moar_freq.get(norm(x.get('moar', '')), 99) <= rare_max
            rank = ns + (0.12 if rare else 0) + (0.04 if x.get('phone') else 0)
            if rank > brank:
                brank, bidx, bname = rank, i, ns
        if bidx is None:
            return None
        best = self.rows[bidx]
        qf = (norm(max(cands, key=lambda c: sim(c, best.get('name', '')))).split() or [''])[0]
        bf = (norm(best.get('name', '')).split() or [''])[0]
        ftok = difflib.SequenceMatcher(None, qf, bf).ratio() >= ftok_th
        ms = sim(moar, best.get('moar', '')) if moar else 0
        rare = ms >= 0.8 and self.moar_freq.get(norm(best.get('moar', '')), 99) <= rare_max
        if ftok and (bname >= th or (bname >= 0.75 and rare)):
            return bidx
        return None

    def match(self, name_candidates, moar='', subj='', th=0.82, ftok_th=0.55, rare_max=4):
        """يطابق الاسم؛ يعيد صف المرجع المطابق أو None."""
        i = self._match_idx(name_candidates, moar=moar, subj=subj, th=th, ftok_th=ftok_th, rare_max=rare_max)
        return self.rows[i] if i is not None else None

    def _weak_evidence(self, sig, cand):
        """قرينة تكفي لقبول صف مرشّح بموقعه التسلسلي (لا للمطابقة الحرة).
        قائمة على الاسم فقط: الموضوع/المعرف الشائعان (نقل، محمد الاسدي…)
        يتشابهان بين الجيران فيقبلان الصف الخطأ — جُرِّب وفشل."""
        nm, cn = sig.get('name', ''), cand.get('name', '')
        if not nm:
            return False
        if sim(nm, cn) >= 0.45:
            return True
        qt, ct = norm(nm).split(), norm(cn).split()
        if qt and ct:
            f = difflib.SequenceMatcher(None, qt[0], ct[0]).ratio()
            l = difflib.SequenceMatcher(None, qt[-1], ct[-1]).ratio()
            if f >= 0.7 and l >= 0.7:
                return True
        return False

    def align(self, signals, th=0.82, ftok_th=0.55, rare_max=4):
        """يطابق قائمة صفوف كاملة مع استغلال التسلسل: صفحات السجل غالباً مقطع
        متتالٍ من صفوف المرجع. بعد المطابقة الحرة نحدّد «مراسي موثوقة»
        (مطابقات تتشارك الإزاحة فهرس المرجع − موقع الصف مع مرساة أخرى)،
        ولكل صف نجرّب إزاحة أقرب مرساة يساراً ثم يميناً — الإزاحة محلية لا
        عامة، حتى لا ينزاح كل شيء إن نقص شخص واحد من المرجع (جُرِّب وفشل).
        signals: قائمة {'name','moar','subj'}. يعيد قائمة صفوف مرجع أو None."""
        strict = [self._match_idx([s.get('name', '')], moar=s.get('moar', ''), subj=s.get('subj', ''),
                                  th=th, ftok_th=ftok_th, rare_max=rare_max) for s in signals]
        hits = [self.rows[i] if i is not None else None for i in strict]
        offs = {k: i - k for k, i in enumerate(strict) if i is not None}
        cnt = Counter(offs.values())
        trusted = {k: o for k, o in offs.items() if cnt[o] >= 2}
        if len(trusted) < 3:
            return hits          # لا نمط تسلسلياً واضحاً — نكتفي بالمطابقة الحرة
        keys = sorted(trusted)
        out = []
        for k, sig in enumerate(signals):
            left = [t for t in keys if t <= k]
            right = [t for t in keys if t >= k]
            offsets_try = []
            if left:
                offsets_try.append(trusted[left[-1]])
            if right and trusted[right[0]] not in offsets_try:
                offsets_try.append(trusted[right[0]])
            chosen = None
            for o in offsets_try:
                j = o + k
                if 0 <= j < len(self.rows) and self._weak_evidence(sig, self.rows[j]):
                    chosen = self.rows[j]   # الموقع التسلسلي يغلب (يصحّح مطابقة تكرار اسمٍ خاطئة)
                    break
            out.append(chosen if chosen is not None else hits[k])
        return out
