# -*- coding: utf-8 -*-
"""تصدير Excel منسّق بهوية النهج + ألوان الثقة + عمود المصدر."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

NHJ_TEAL = '19BBBD'
CONF_FILL = {
    'ref':    'C6EFCE',   # أخضر: مؤكد من المرجع
    'word':   'C6EFCE',   # أخضر: مؤكد من ملف Word (بعد التأكيد)
    'agree':  'DDEBF7',   # أزرق: ثقة عالية
    'review': 'FFF2CC',   # أصفر: يحتاج مراجعة
    'phone_unconf': 'FCE4D6',  # برتقالي: هاتف غير مؤكد
}


def export(out_path, headers, rows, colors=None, source_col=True):
    """rows: list[dict header->value]. colors: {(row_idx0, header): conf_key}."""
    colors = colors or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'سجل الصادر'
    ws.sheet_view.rightToLeft = True

    out_headers = list(headers) + (['المصدر'] if source_col else [])
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfont = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
    cfont = Font(name='Segoe UI', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    hfill = PatternFill('solid', fgColor=NHJ_TEAL)

    for c, h in enumerate(out_headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = hfill; cell.font = hfont; cell.alignment = center; cell.border = border

    num_like = {'رقم الكتاب', 'رقم', 'التسلسل'}
    for i, row in enumerate(rows):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(i + 2, c, row.get(h, ''))
            cell.font = cfont; cell.border = border
            cell.alignment = center if h in num_like else right
            key = colors.get((i, h))
            if key and key in CONF_FILL:
                cell.fill = PatternFill('solid', fgColor=CONF_FILL[key])
        if source_col:
            src = row.get('_source', '')
            sc = ws.cell(i + 2, len(out_headers), src)
            sc.font = cfont; sc.border = border; sc.alignment = center

    widths = []
    for h in out_headers:
        widths.append(11 if h in num_like else (12 if h == 'المصدر' else (26 if 'موضوع' in h or 'جهة' in h else 20)))
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 28
    wb.save(out_path)
    return out_path
