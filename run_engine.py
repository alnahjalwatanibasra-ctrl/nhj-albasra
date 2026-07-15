# -*- coding: utf-8 -*-
"""مشغّل تجريبي للمرحلة 1: يشغّل المحرّك على صور الاختبار ويقارن بالمرجع الذهبي."""
import sys, os, unicodedata
sys.stdout.reconfigure(encoding='utf-8')
from core import pipeline, config
from core.corrections import to_western_digits
import openpyxl

SER = r'C:\Users\ABR ALSHARQ\Desktop\ser'
IMAGES = [os.path.join(SER, f) for f in [
    'WhatsApp Image 2026-07-14 at 4.03.38 PM.jpeg',
    'WhatsApp Image 2026-07-14 at 4.03.38 PM (1).jpeg',
    'WhatsApp Image 2026-07-14 at 4.03.38 PM (2).jpeg']]
REF = os.path.join(SER, 'انجازات النهج.xlsx')
GOLD = os.path.join(SER, 'سجل_الصادر.xlsx')   # الملف المصحّح 100%
OUT = os.path.join(SER, 'مخرجات_البرنامج.xlsx')

def norm(s):
    s='' if s is None else str(s)
    for a,b in [('أ','ا'),('إ','ا'),('آ','ا'),('ى','ي'),('ة','ه'),('ؤ','و'),('ئ','ي'),('ـ','')]: s=s.replace(a,b)
    return ' '.join(''.join(c for c in s if not unicodedata.combining(c)).split()).strip()

print('>> تشغيل المحرّك على 3 صور ...')
res = pipeline.run(IMAGES, REF, prev_register_path=None,
                   progress=lambda m: print('  ', m),
                   cache_path=os.path.join(os.path.dirname(__file__), 'extract_cache.json'))
print('العناوين:', res['headers'])
print('أعمدة المرجع المكتشفة:', res['ref_columns'])
print('عدد الصفوف:', len(res['rows']))
print('صفوف مطابقة للمرجع:', sum(res['matched']))

from core.excel_export import export
export(OUT, res['headers'], res['rows'], colors=res['colors'], source_col=False)
print('حُفظت المخرجات:', OUT)

# مقارنة بالمرجع الذهبي (بالموضع) — كل الأعمدة
wb=openpyxl.load_workbook(GOLD, data_only=True); ws=wb.active
gh=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
def col(name): return gh.index(name)+1 if name in gh else None
GCOLS={'رقم':col('رقم الكتاب'),'اسم':col('اسم صاحب الكتاب'),'موضوع':col('موضوع الكتاب'),
       'هاتف':col('رقم الهاتف'),'تاريخ':col('تاريخ الكتاب'),'معرف':col('المعرف'),
       'جهة':col('الجهة المرسل اليها')}
gold=[]
for r in range(2,ws.max_row+1):
    if not ws.cell(r,GCOLS['اسم']).value: continue
    gold.append({k:ws.cell(r,c).value for k,c in GCOLS.items() if c})
OCOLS={'رقم':'رقم الكتاب','اسم':'اسم صاحب الكتاب','موضوع':'موضوع الكتاب','هاتف':'رقم الهاتف',
       'تاريخ':'تاريخ الكتاب','معرف':'المعرف','جهة':'الجهة المرسل اليها'}
def gdate(v):
    if v is None: return ''
    if hasattr(v,'year'): return '%d/%d/%d'%(v.month,v.day,v.year)
    return str(v).strip()
def gphone(v):
    if v in (None,'','لا يوجد'): return 'لا يوجد' if v=='لا يوجد' else ''
    return to_western_digits(v).lstrip('0')
out=res['rows']
n=min(len(out),len(gold)); acc={k:[0,0] for k in OCOLS}; diffs=[]
for i in range(n):
    g=gold[i]; o=out[i]
    for k in OCOLS:
        gv, ov = g.get(k), o.get(OCOLS[k],'')
        if k=='رقم': gv2, ov2 = to_western_digits(gv), to_western_digits(ov)
        elif k=='هاتف':
            gv2 = gphone(gv)
            ov2 = 'لا يوجد' if str(ov).strip()=='لا يوجد' else to_western_digits(ov).lstrip('0')
        elif k=='تاريخ': gv2, ov2 = gdate(gv), to_western_digits(str(ov)) and gdate(str(ov)) or ''
        else: gv2, ov2 = norm(gv), norm(ov)
        if k=='تاريخ':
            gv2 = gdate(gv); ov2 = str(ov).translate(str.maketrans('٠١٢٣٤٥٦٧٨٩','0123456789')).strip()
        if not str(gv2): continue
        acc[k][1]+=1
        if str(gv2)==str(ov2): acc[k][0]+=1
        else: diffs.append((i,k,gv2,ov2))
print('\n=== دقة المحرّك الكامل مقابل الذهبي (كل الأعمدة) ===')
for k,(ok,tot) in acc.items():
    print(f'  {k}: {ok}/{tot} ({round(100*ok/tot) if tot else 0}%)')
if diffs:
    print('\n-- الفروقات --')
    for i,k,gv,ov in diffs: print(f'  صف{i+1} [{k}] ذهبي={gv!r} مخرج={ov!r}')
